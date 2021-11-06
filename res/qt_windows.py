import getpass
import itertools
import os
import socket
from collections import defaultdict
from copy import copy
from datetime import datetime
from itertools import cycle
from math import inf
from random import randint
from typing import Optional, Union, List, Iterable, Tuple

import openpyxl as opx
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter as letter
from PyQt5 import QtCore, QtGui, QtWidgets, uic
from PyQt5.QtCore import QSize, QPoint
from PyQt5.QtGui import QBrush, QImage, QPalette, QPixmap, QDragLeaveEvent
from PyQt5.QtWidgets import (
    QCheckBox,
    QComboBox,
    QGridLayout,
    QLabel,
    QLineEdit,
    QSpacerItem,
    QListWidget, QPushButton, QRadioButton, QLayout, QButtonGroup, QWidget, QListWidgetItem, QTextEdit,
)

from res.additional_classes import MyQLabel, MyQGridLayout
from res.additional_funcs import set_qt_stile

from res.edit_db_windows import EditDataWindow, EditMaterialWindow
from res.material_classes import Material, Receipt, Profile, ReceiptData


class MyMainWindow(QtWidgets.QMainWindow, uic.loadUiType("windows/Main_window.ui")[0]):
    saved_receipt_listWidget: "SavedReceiptWidget"

    lineEdit_name_a: QLineEdit  # Строка с названием Компонента А
    lineEdit_name_b: QLineEdit  # Строка с названием Компонента Б
    comment_textEdit_a: QTextEdit  # Строка с комментарием Компонента А
    comment_textEdit_b: QTextEdit  # Строка с комментарием Компонента Б

    final_a: QLabel  # Строка "ИТОГО" в конце рецептуры
    final_b: QLabel  # Строка "ИТОГО" в конце рецептуры
    final_a_numb_label: QLabel  # Сумма процентов в рецептуре А
    final_b_numb_label: QLabel  # Сумма процентов в рецептуре Б
    final_mass_a: QLineEdit  # Масса загрузки компонента А
    final_mass_b: QLineEdit  # Масса загрузки компонента Б
    mass_radio_used_group_a: QButtonGroup  # QButtonGroup для объединения QRadioButton
    mass_radio_used_group_b: QButtonGroup  # QButtonGroup для объединения QRadioButton
    default_radio_a: QRadioButton  # QRadioButton для отключения логики фиксированной массы одного компонента
    default_radio_b: QRadioButton  # QRadioButton для отключения логики фиксированной массы одного компонента

    cur_mass_radio_used_a: int  # Текущий индекс mass_radio_used_group_a
    cur_mass_radio_used_b: int  # Текущий индекс mass_radio_used_group_b
    remember_mass_a: float  # Переменная для запоминания массы
    remember_mass_b: float  # Переменная для запоминания массы

    adw_a: "AcceptDropWidget"  # Область для перетаскивания рецептуры для установки в компонент А
    adw_b: "AcceptDropWidget"  # Область для перетаскивания рецептуры для установки в компонент Б
    receipt_bin: "AcceptDropWidget"  # Область для перетаскивания рецептуры для удаления

    all_receipts: List[ReceiptData]  # Список всех рецептур в профиле

    def __init__(self, profile: Profile, init_class, debug=False):
        super(MyMainWindow, self).__init__()
        self.setupUi(self)

        self.profile = profile
        self.debug_flag = debug
        self.init_class = init_class


        # TODO Вынести изменения шрифта в init_class, чтобы установка происходила через него и сразу всем
        self.all_labels = [
            self.mass_ratio_label,
            self.tg_main_label,
            self.mass_ratio_label_2,
            self.label_3,
            self.label_4,
            self.label_5,
            self.label_6,
            self.label_7,
            self.eew_label,
            self.ahew_label,
            self.extra_ew_label,
            self.sintez_pair_label,
            self.debug_string,
            self.lineEdit_name_a,
            self.lineEdit_name_b,
            self.tg_cor_label,
            self.tg_extra_label,
            self.extra_ratio_line,
        ]
        self.all_big_labels = [self.label, self.label_2]
        self.font_size = 9
        self.font_size_big = 5

        # QSpacerItem в gridLayout для подпирания строк снизу
        self.gridLayout_a.addItem(QSpacerItem(100, 100), 100, 0, 100, 20)
        self.gridLayout_b.addItem(QSpacerItem(100, 100), 100, 0, 100, 20)

        with open("style.css", "r") as f:
            self.style, self.style_combobox, self.style_red_but = f.read().split(
                "$split$"
            )

        set_qt_stile("style.css", self)
        self.set_inner_style()

        self.types_of_items = []
        self.update_list_of_material_types()
        self.list_of_material_names = {}
        self.update_list_of_material_names()

        # ================ Инициализация переменных для строки ИТОГО в рецептуре ==============
        self.setup_final_lines()

        # ========================  Контейнеры для хранения данных ========================

        # QComboBox с типами материалов
        self.material_a_types: List[QComboBox] = []
        self.material_b_types: List[QComboBox] = []
        # QComboBox с названиями материалов
        self.material_comboboxes_a: List[QComboBox] = []
        self.material_comboboxes_b: List[QComboBox] = []
        # QLineEdit с процентами материалов
        self.material_percent_lines_a: List[QLineEdit] = []
        self.material_percent_lines_b: List[QLineEdit] = []
        # QCheckBox для логики фиксации процентов при нормировании
        self.lock_checkboxies_a: List[QCheckBox] = []
        self.lock_checkboxies_b: List[QCheckBox] = []

        # QLineEdit с массами загрузки
        self.mass_lines_a: List[QLineEdit] = []
        self.mass_lines_b: List[QLineEdit] = []
        # QRadioButton для логики работы с массами загрузки
        self.mass_radio_list_a: List[QRadioButton] = []
        self.mass_radio_list_b: List[QRadioButton] = []
        self.mass_radio_used_group_a = QButtonGroup()
        self.mass_radio_used_group_b = QButtonGroup()
        self.default_radio_a = QRadioButton()
        self.default_radio_b = QRadioButton()
        self.mass_radio_used_group_a.addButton(self.default_radio_a, 0)
        self.mass_radio_used_group_b.addButton(self.default_radio_b, 0)
        # Объекты класса Material
        self.material_list_a: List[Material] = []
        self.material_list_b: List[Material] = []
        # Рецептуры, в которых вся логика рецептур
        self.receipt_a: Optional[Receipt] = None
        self.receipt_b: Optional[Receipt] = None

        self.receipt_data_a: Optional[ReceiptData] = None
        self.receipt_data_b: Optional[ReceiptData] = None

        # ======================== Переменные для окон =====================================

        self.a_receipt_window: Optional[SintezWindow] = None
        self.b_receipt_window: Optional[SintezWindow] = None

        self.pair_react_window: Optional[PairReactWindow] = None

        # ======= Подключаем кнопки =============================================
        self.setup_connect_buttons()
        # ======= Настройка логики перетаскивания рецептуры ======================
        self.setup_receipts_drag_logic()

        # ======================== тесты по зеленым квадратикам ============================

        self.inf_window = None
        self.warring_grid = MyQGridLayout(self.centralwidget)

    # ================== Функции для инициализации данных ============================

    def setup_final_lines(self):
        """
        Инициализация переменных для последних строк рецептур
        """
        self.final_a = None
        self.final_b = None
        self.final_a_numb_label = None
        self.final_b_numb_label = None
        self.final_mass_a: QLineEdit = None
        self.final_mass_b: QLineEdit = None
        self.cur_mass_radio_used_a = None
        self.cur_mass_radio_used_b = None
        self.remember_mass_a = 0.0
        self.remember_mass_b = 0.0
        self.hide_top("A")
        self.hide_top("B")

    def setup_connect_buttons(self):
        """
        Подключение всех кнопок окна
        """
        # ======================== Подключаем кнопки =======================================
        self.add_A_but.clicked.connect(self.add_a_line)
        self.add_B_but.clicked.connect(self.add_b_line)
        self.del_A_but.clicked.connect(self.del_a_line)
        self.del_B_but.clicked.connect(self.del_b_line)

        self.normalise_A.clicked.connect(self.normalise_func("A"))
        self.normalise_B.clicked.connect(self.normalise_func("B"))

        self.a_receipt_but.clicked.connect(self.add_receipt_window("A"))
        self.b_receipt_but.clicked.connect(self.add_receipt_window("B"))

        # ====================================== Кнопки меню =======================================
        self.menu_sintez_edit.triggered.connect(self.add_pair_react_window)
        self.menu_prof_edit.triggered.connect(self.add_profile_edit_window)
        self.menu_add_mat.triggered.connect(self.add_material_window)

        self.excel_save_A.triggered.connect(lambda: self.export_to_excel("A"))
        self.excel_save_B.triggered.connect(lambda: self.export_to_excel("B"))
        self.excel_save_all.triggered.connect(lambda: self.export_to_excel("AB"))

        self.save_a.triggered.connect(lambda: self.save_receipt("A"))
        self.save_b.triggered.connect(lambda: self.save_receipt("B"))

        self.menu_change_profile.triggered.connect(self.change_profile)

        # ====================================== Кнопки дебага =======================================
        self.debug_but.clicked.connect(self.debug)
        self.update_but.clicked.connect(self.debug_2)

    def setup_receipts_drag_logic(self):
        """
        Подключает логику работы с рецептурами
        :return:
        """

        self.adw_a = AcceptDropWidget(self, 'load', "A")
        self.adw_a.move(QPoint(405, 35))
        self.adw_a.resize(QSize(100, 40))
        self.adw_a.show()

        pixmap = QPixmap("icons/set_receipt.png")
        self.adw_a.setScaledContents(True)
        self.adw_a.setPixmap(pixmap)

        self.adw_b = AcceptDropWidget(self, 'load', "B")
        self.adw_b.move(QPoint(785, 35))
        self.adw_b.resize(QSize(100, 40))
        self.adw_b.show()

        self.adw_b.setScaledContents(True)
        self.adw_b.setPixmap(pixmap)

        self.receipt_bin = AcceptDropWidget(self, "delete")
        self.receipt_bin.move(QPoint(1080, 35))
        self.receipt_bin.resize(QSize(30, 30))
        self.receipt_bin.show()

        pixmap = QPixmap("icons/bin.png")
        self.receipt_bin.setScaledContents(True)
        self.receipt_bin.setPixmap(pixmap)

        self.saved_receipt_listWidget = SavedReceiptWidget(self)
        self.all_receipts = sorted(self.profile.orm_db.get_profile_receipts(self.profile),
                                   key=lambda x: x.date, reverse=True)
        for index, receipt in enumerate(self.all_receipts):
            self.saved_receipt_listWidget.addItem(str(receipt.name))
            self.saved_receipt_listWidget.item(index).setToolTip(str(receipt.comment))

    def change_profile(self):
        user_name: str = getpass.getuser()
        computer_name: str = socket.gethostname()
        self.profile.orm_db.remove_computer_from_db(user_name, computer_name)
        self.init_class.choose_profile()
        self.close()

    def debug(self) -> None:

        self.add_a_line()
        self.add_a_line()
        self.material_a_types[0].setCurrentIndex(2)
        self.material_a_types[1].setCurrentIndex(2)
        self.material_comboboxes_a[0].setCurrentIndex(3)
        self.material_comboboxes_a[1].setCurrentIndex(4)
        self.material_percent_lines_a[0].setText("50.00")
        self.material_list_a[0].percent = 50
        self.material_percent_lines_a[1].setText("50.00")
        self.material_list_a[1].percent = 50

        # self.normalise_func('A')
        # self.to_float('A')
        self.normalise_func("A")

        self.add_b_line()
        self.add_b_line()
        # self.material_b_types[0].setCurrentIndex(2)
        # self.material_b_types[1].setCurrentIndex(2)
        self.material_comboboxes_b[0].setCurrentIndex(4)
        self.material_comboboxes_b[1].setCurrentIndex(3)
        self.material_percent_lines_b[0].setText("50.00")
        self.material_list_b[0].percent = 50
        self.material_percent_lines_b[1].setText("50.00")
        self.material_list_b[1].percent = 50

    def debug_2(self) -> None:
        # self.create_warring()

        # self.save_receipt('A')
        self.export_to_excel('AB')
        print('debug')

    def hide_top(self, component: str) -> None:
        """
        Прячем верх рецептуры, когда нет компонентов
        :param component:
        :return:
        """
        if component == "A":
            self.label_3.hide()
            self.label_5.hide()
            self.normalise_A.hide()
            self.label_lock_a.hide()
            self.normalise_mass_A.hide()

        elif component == "B":
            self.label_4.hide()
            self.label_6.hide()
            self.normalise_B.hide()
            self.label_lock_b.hide()
            self.normalise_mass_B.hide()

    def add_line(self, component: str) -> None:
        """
        Добавляет строку компонента в рецептуру
        :param component: А или Б
        :return:
        """
        final_label = QLabel("Итого")
        final_label.setStyleSheet(self.style)
        final_label.setFont((QtGui.QFont("Times New Roman", self.font_size)))
        final_label_numb = QLabel("0.00")
        final_label_numb.setStyleSheet(self.style)
        final_label_numb.setFont((QtGui.QFont("Times New Roman", self.font_size)))

        final_label_numb.setTextInteractionFlags(
            QtCore.Qt.LinksAccessibleByMouse
            | QtCore.Qt.TextSelectableByKeyboard
            | QtCore.Qt.TextSelectableByMouse
        )

        final_mass_line = QLineEdit()
        final_mass_line.setText("0.00")
        final_mass_line.setFixedWidth(50)
        final_mass_line.setFont((QtGui.QFont("Times New Roman", self.font_size)))
        final_mass_line.setStyleSheet("QLineEdit{background : lightblue;}")
        final_mass_line.editingFinished.connect(lambda: (self.to_float(component),
                                                         self.update_mass_in_component_lines(component)))

        if component == "A":
            items_type = self.material_a_types
            items = self.material_comboboxes_a
            items_lines = self.material_percent_lines_a
            grid = self.gridLayout_a
            lock_checkboxes = self.lock_checkboxies_a
            if self.final_a:
                final_mass_line.setText(self.final_mass_a.text())
                self.final_a.deleteLater()
                self.final_a_numb_label.deleteLater()
                self.final_mass_a.deleteLater()

            self.final_a = final_label
            self.final_a_numb_label = final_label_numb
            self.final_mass_a = final_mass_line
            receipt = self.receipt_a
            mass_lines = self.mass_lines_a
            radio_lines = self.mass_radio_list_a
            group_button = self.mass_radio_used_group_a

        elif component == "B":
            items_type = self.material_b_types
            items = self.material_comboboxes_b
            items_lines = self.material_percent_lines_b
            grid = self.gridLayout_b

            lock_checkboxes = self.lock_checkboxies_b
            if self.final_b:
                final_mass_line.setText(self.final_mass_b.text())
                self.final_b.deleteLater()
                self.final_b_numb_label.deleteLater()
                self.final_mass_b.deleteLater()
            self.final_b = final_label
            self.final_b_numb_label = final_label_numb
            self.final_mass_b = final_mass_line
            receipt = self.receipt_b
            mass_lines = self.mass_lines_b
            radio_lines = self.mass_radio_list_b
            group_button = self.mass_radio_used_group_b
        else:
            return None

        self.show_top(component)

        row_count = len(items)

        material_combobox = QComboBox()
        material_combobox.setFixedWidth(120)
        material_combobox.setFixedHeight(20)
        material_combobox.setStyleSheet(self.style_combobox)
        material_combobox.setFont((QtGui.QFont("Times New Roman", self.font_size)))

        materia_typel_combobox = QComboBox()
        materia_typel_combobox.setFixedWidth(60)
        materia_typel_combobox.currentIndexChanged.connect(
            self.change_list_of_materials(material_combobox, materia_typel_combobox)
        )

        materia_typel_combobox.addItems(self.types_of_items)
        materia_typel_combobox.setFixedHeight(20)
        materia_typel_combobox.setFont(QtGui.QFont("Times New Roman", self.font_size))
        materia_typel_combobox.setStyleSheet(self.style_combobox)

        percent_line = QLineEdit()
        percent_line.setText("0.00")
        percent_line.setFixedWidth(45)
        percent_line.setFont((QtGui.QFont("Times New Roman", self.font_size)))
        percent_line.editingFinished.connect(lambda: self.to_float(component))

        mass_line = QLineEdit()
        mass_line.setText("0.00")
        mass_line.setFixedWidth(50)
        mass_line.setFont((QtGui.QFont("Times New Roman", self.font_size)))
        mass_line.editingFinished.connect(lambda: self.to_float(component))

        radio = QRadioButton()
        radio.resize(20, 20)
        radio.clicked.connect(self.update_mass_radio(row_count, component))
        group_button.addButton(radio, row_count + 1)

        material = Material(materia_typel_combobox.currentText(), material_combobox.currentIndex(),
                            self.profile, receipt)
        # Подключение функций, которые передают параметры в Material при изменении
        percent_line.editingFinished.connect(
            self.change_percent_material(material, percent_line)
        )

        material_combobox.currentIndexChanged.connect(
            self.change_type_name_material(
                material, materia_typel_combobox, material_combobox
            )
        )
        check = QCheckBox()
        check.setFixedWidth(22)
        lock_checkboxes.append(check)
        items_type.append(materia_typel_combobox)
        items.append(material_combobox)
        items_lines.append(percent_line)
        mass_lines.append(mass_line)
        radio_lines.append(radio)
        grid: QGridLayout
        grid.addWidget(materia_typel_combobox, row_count + 1, 0)
        grid.addWidget(material_combobox, row_count + 1, 1)
        grid.addWidget(percent_line, row_count + 1, 2)
        grid.addWidget(check, row_count + 1, 3)
        grid.addWidget(mass_line, row_count + 1, 4)
        grid.addWidget(radio, row_count + 1, 5)

        grid.addWidget(final_label, row_count + 2, 1, alignment=QtCore.Qt.AlignRight)
        grid.addWidget(final_label_numb, row_count + 2, 2)
        grid.addWidget(final_mass_line, row_count + 2, 4)

        receipt.count_sum()

    def add_a_line(self) -> None:
        self.add_line("A")

    def add_b_line(self) -> None:
        self.add_line("B")

    # Отображает шапку рецептуры, когда есть компоненты
    def show_top(self, component: str):
        """Отображает шапку рецептуры, когда есть компоненты"""
        if component == "A":
            self.label_3.show()
            self.label_5.show()
            self.normalise_A.show()
            self.normalise_mass_A.show()
            self.label_lock_a.show()

        if component == "B":
            self.label_4.show()
            self.label_6.show()
            self.normalise_B.show()
            self.normalise_mass_B.show()
            self.label_lock_b.show()

    # Удаляет последнюю строку в рецептуре
    def del_line(self, component: str) -> None:
        """Удаляет последнюю строку в рецептуре"""

        final_mass = "0.00"
        if component == "A":
            if len(self.material_list_a) == 0:
                return None
            items_type = self.material_a_types
            items = self.material_comboboxes_a
            items_lines = self.material_percent_lines_a
            mass_lines = self.mass_lines_a
            mass_radio = self.mass_radio_list_a
            grid = self.gridLayout_a
            lock_check_boxes = self.lock_checkboxies_a
            if self.final_a:
                final_mass = self.final_mass_a.text()
                self.final_a.deleteLater()
                self.final_a = None
                self.final_a_numb_label.deleteLater()
                self.final_a_numb_label = None
                self.final_mass_a.deleteLater()
                self.final_mass_a = None

            self.receipt_a.remove_material()

        elif component == "B":
            if len(self.material_list_b) == 0:
                return None
            items_type = self.material_b_types
            items = self.material_comboboxes_b
            items_lines = self.material_percent_lines_b
            mass_lines = self.mass_lines_b
            mass_radio = self.mass_radio_list_b
            grid = self.gridLayout_b
            lock_check_boxes = self.lock_checkboxies_b
            if self.final_b:
                final_mass = self.final_mass_b.text()
                self.final_b.deleteLater()
                self.final_b = None
                self.final_b_numb_label.deleteLater()
                self.final_b_numb_label = None
                self.final_mass_b.deleteLater()
                self.final_mass_b = None
            self.receipt_b.remove_material()

        else:
            return None

        if items:
            items.pop(-1).deleteLater()
            items_lines.pop(-1).deleteLater()
            items_type.pop(-1).deleteLater()
            lock_check_boxes.pop(-1).deleteLater()
            mass_lines.pop(-1).deleteLater()
            mass_radio.pop(-1).deleteLater()

            if items:
                final_label = QLabel("Итого")
                final_label.setStyleSheet(self.style)
                final_label.setFont((QtGui.QFont("Times New Roman", self.font_size)))
                final_label_numb = QLabel()
                final_label_numb.setStyleSheet(self.style)
                final_label_numb.setFont(
                    (QtGui.QFont("Times New Roman", self.font_size))
                )

                final_mass_line = QLineEdit()
                final_mass_line.setText(final_mass)
                final_mass_line.setFixedWidth(50)
                final_mass_line.setFont((QtGui.QFont("Times New Roman", self.font_size)))
                final_mass_line.setStyleSheet("QLineEdit{background : lightblue;}")
                final_mass_line.editingFinished.connect(lambda: self.to_float(component))

                grid: QGridLayout
                row_count = grid.count()
                grid.addWidget(final_label, row_count + 1, 1, alignment=QtCore.Qt.AlignRight)
                grid.addWidget(final_label_numb, row_count + 1, 2)
                grid.addWidget(final_mass_line, row_count + 1, 4)
                if component == "A":
                    self.final_a = final_label
                    self.final_a_numb_label = final_label_numb
                    self.final_mass_a = final_mass_line
                    self.receipt_a.count_sum()
                elif component == "B":
                    self.final_b = final_label
                    self.final_b_numb_label = final_label_numb
                    self.final_mass_b = final_mass_line
                    self.receipt_b.count_sum()
            else:
                self.hide_top(component)

    def del_a_line(self) -> None:
        self.del_line("A")

    def del_b_line(self) -> None:
        self.del_line("B")

    def set_sum(self, percent: float, component: str) -> None:
        """
        Устанавливает значение в строк Итого в конце рецептуры
        Если значение != 100, то красит красным
        :param percent: Сумма процентов рецептуры
        :param component: рецептура А или Б
        """
        if component == "A":
            self.final_a_numb_label.setText(f"{percent:.{2}f}")
            if percent != 100:
                self.final_a_numb_label.setStyleSheet("QLabel { color: red}")
                self.a_receipt_but.setStyleSheet(self.style_red_but)
            else:
                self.final_a_numb_label.setStyleSheet("QLabel { color: green}")
                self.a_receipt_but.setStyleSheet(self.style)
        elif component == "B":
            self.final_b_numb_label.setText(f"{percent:.{2}f}")
            if percent != 100:
                self.final_b_numb_label.setStyleSheet("QLabel { color: red}")
                self.b_receipt_but.setStyleSheet(self.style_red_but)
            else:
                self.final_b_numb_label.setStyleSheet("QLabel { color: green}")
                self.b_receipt_but.setStyleSheet(self.style)

    def set_ew(self, component: str, ew: Union[float, None]) -> None:
        """
        Устанавливает EW соответствующей рецептуре
        :param component: рецептура А или Б
        :param ew: Эквивалентный вес рецептуры
        """
        if component == "A":
            if ew is None:
                self.eew_label.setText(f"No EW")
            elif ew > 0:
                self.eew_label.setText(f"EEW = {round(ew, 2)}")
            elif ew < 0:
                self.eew_label.setText(f"AHEW = {-round(ew, 2)}")

        if component == "B":
            if ew is None:
                self.ahew_label.setText(f"No EW")
            elif ew > 0:
                self.ahew_label.setText(f"EEW = {round(ew, 2)}")
            elif ew < 0:
                self.ahew_label.setText(f"AHEW = {-round(ew, 2)}")

    def set_mass_ratio(self, value: Union[float, None]) -> None:
        """
        Устанавливает Массовое соотношение в соответствующий лейбл
        :param value: Значение соотношения.
        """
        if value is None:
            self.mass_ratio_label.setText(f"Продукты не реагируют")
        elif value >= 1:
            numb_a = round(value, 2)
            numb_b = 1
            self.mass_ratio_label.setText(
                f"Соотношение по массе:\n\t{numb_a} : {numb_b}"
            )
        elif 0 < value < 1:
            numb_a = 1
            numb_b = round(1 / value, 2)
            self.mass_ratio_label.setText(
                f"Соотношение по массе:\n\t{numb_a} : {numb_b}"
            )

    def set_tg(self, value) -> None:
        """
        Устанавливает базовое стеклование в соответствующий лейбл
        :param value: Значение базового стеклования
        """
        if value is None:
            self.tg_main_label.setText(f"Стеклование базовое:\n\tотсутствует")
        else:
            self.tg_main_label.setText(f"Стеклование базовое:\n\t{round(value, 1)}°C")

    def set_tg_inf(self, value) -> None:
        """
        Устанавливает базовое стеклование в соответствующий лейбл
        :param value: Значение базового стеклования
        """
        if value is None:
            self.tg_cor_label.setText(f"Стеклование с коррекцией:\n\tотсутствует")
        else:
            self.tg_cor_label.setText(f"Стеклование с коррекцией:\n\t{round(value, 1)}°C")

    # ========================= Окна ===========================

    def add_receipt_window(self, component: str) -> callable:
        """
        Функция для вызова окна редактирования рецептуры. Обёрнута замыканием для вызова триггером кнопки.
        :param component: рецептура А или Б
        :return:
        """

        def wrapper():
            if component == "A":
                if self.final_a_numb_label is not None:
                    if self.final_a_numb_label.text() == "100.00":
                        if not self.a_receipt_window:
                            self.a_receipt_window = SintezWindow(self, "A")
                        if not self.debug_flag:
                            self.a_receipt_window.show()
                        self.disable_receipt("A")
            elif component == "B":
                if self.final_b_numb_label is not None:
                    if self.final_b_numb_label.text() == "100.00":
                        if not self.b_receipt_window:
                            self.b_receipt_window = SintezWindow(self, "B")
                        if not self.debug_flag:
                            self.b_receipt_window.show()
                        self.disable_receipt("B")
            else:
                return None

        return wrapper

    def add_pair_react_window(self) -> None:
        """
        Функция для вызова окна редактирования пар
        """
        if not self.pair_react_window:
            self.pair_react_window = PairReactWindow(
                self, self.receipt_a, self.receipt_b
            )
        if not self.debug_flag:
            self.pair_react_window.show()

    def add_profile_edit_window(self):
        self.profile_edit_window = EditDataWindow(self, self.profile)
        self.profile_edit_window.show()
        self.close()

    def add_material_window(self):
        self.material_window = EditMaterialWindow(self, self.profile)
        self.close()
        self.material_window.show()

    # =========================  ===========================

    @staticmethod
    def isfloat(value) -> bool:
        """Проверяет число на возможность приведения к float"""
        try:
            float(value)
            return True
        except ValueError:
            return False

    # Приводит все проценты в рецептуре к типу float и считает +-*/ если есть в строке
    def to_float(self, component: str) -> None:
        """Приводит все проценты в рецептуре к типу float и считает +-*/ если есть в строке"""
        if component == "A":
            percent_lines = self.material_percent_lines_a
            mass_lines = self.mass_lines_a
            final_mass = [self.final_mass_a]
        elif component == "B":
            percent_lines = self.material_percent_lines_b
            mass_lines = self.mass_lines_b
            final_mass = [self.final_mass_b]
        else:
            return None
        numb: Union[str, float]
        for widget in percent_lines + mass_lines + final_mass:
            numb = widget.text().replace(",", ".")
            if len(numb.split("+")) > 1:
                split_numb = numb.split("+")
                if all([i for i in map(self.isfloat, split_numb)]):
                    numb = sum(map(float, split_numb))
                elif self.isfloat(split_numb[0]):
                    numb = float(split_numb[0])
            elif len(numb.split("-")) > 1:
                split_numb = numb.split("-")
                if all([i for i in map(self.isfloat, split_numb)]):
                    numb = float(split_numb[0]) - float(split_numb[1])
                elif self.isfloat(split_numb[0]):
                    numb = float(split_numb[0])
            elif len(numb.split("*")) > 1:
                split_numb = numb.split("*")
                if all([i for i in map(self.isfloat, split_numb)]):
                    numb = float(split_numb[0]) * float(split_numb[1])
                elif self.isfloat(split_numb[0]):
                    numb = float(split_numb[0])
            elif len(numb.split("/")) > 1:
                split_numb = numb.split("/")
                if all([i for i in map(self.isfloat, split_numb)]):
                    numb = float(split_numb[0]) / float(split_numb[1])
                elif self.isfloat(split_numb[0]):
                    numb = float(split_numb[0])
            elif len(numb.split("\\")) > 1:
                split_numb = numb.split("\\")
                if all([i for i in map(self.isfloat, split_numb)]):
                    numb = float(split_numb[0]) / float(split_numb[1])
                elif self.isfloat(split_numb[0]):
                    numb = float(split_numb[0])

            if not self.isfloat(numb):
                numb = 0

            if float(numb) < 0:
                numb = 0
            widget.setText(f"{float(numb):.{2}f}")

    # Меняет список сырья при смене типа в рецептуре
    def change_list_of_materials(
            self, material_combobox: QComboBox, material_type: QComboBox
    ) -> callable:
        """Меняет список сырья при смене типа в рецептуре
        Обёрнута в замыкание для вызова триггером"""

        def wrapper():
            material_combobox.clear()
            material_combobox.addItems(
                self.list_of_material_names[material_type.currentText()]
            )

        return wrapper

    def update_mass_radio(self, current_line: int, component: str) -> callable:
        """
        Обрабатывает смену QRadioButton для строк загрузки
        :param current_line:
        :param component:
        :return:
        """

        def wrapper():
            if component == 'A':
                default = self.mass_radio_used_group_a.button(0)
                if self.cur_mass_radio_used_a == current_line:
                    self.cur_mass_radio_used_a = None
                    default.setChecked(True)
                else:
                    self.cur_mass_radio_used_a = current_line

            elif component == "B":
                default = self.mass_radio_used_group_b.button(0)
                if self.cur_mass_radio_used_b == current_line:
                    self.cur_mass_radio_used_b = None
                    default.setChecked(True)
                else:
                    self.cur_mass_radio_used_b = current_line

        return wrapper

    def update_mass(self):
        self.update_mass_in_component_lines('A')
        self.update_mass_in_component_lines('B')

    def update_mass_in_component_lines(self, component: str):
        if component == "A":
            current_index = self.cur_mass_radio_used_a
            mass_lines = self.mass_lines_a
            material_lines = self.material_list_a
            if current_index is not None:
                mass = float(mass_lines[current_index].text())
                percent = material_lines[current_index].percent
                if mass == 0:
                    mass = self.remember_mass_a
                    mass_lines[current_index].setText(str(mass))
                if percent != 0:
                    self.final_mass_a.setText(str(round(mass / percent * 100, 2)))

                else:
                    self.remember_mass_a = mass
                    mass_lines[current_index].setText('0.00')
            mass = float(self.final_mass_a.text())

        elif component == "B":
            current_index = self.cur_mass_radio_used_b
            mass_lines = self.mass_lines_b
            material_lines = self.material_list_b
            if current_index is not None:
                mass = float(mass_lines[current_index].text())
                percent = material_lines[current_index].percent
                if mass == 0:
                    mass = float(self.final_mass_a.text())
                    mass_lines[current_index].setText(str(mass))
                if percent != 0:
                    self.final_mass_b.setText(str(round(mass / percent * 100, 2)))
                else:
                    mass_lines[current_index].setText('0.00')
            mass = float(self.final_mass_b.text())
        else:
            return None

        for index, mass_line in enumerate(mass_lines):
            if index == current_index:
                continue
            percent = material_lines[index].percent
            numb = round(percent * mass / 100, 2)
            # setText(f"{float(numb):.{2}f}")
            mass_line.setText(f"{float(numb):.{2}f}")

    def update_list_of_material_names(self):
        """
        Обновляет список материалов в окне
        :return:
        """
        self.list_of_material_names = {
            mat_type: self.profile.get_mat_names_by_type(mat_type)
            for mat_type in self.types_of_items
        }
        # TODO Продумать, как добавить материалы в комбобоксы, и чтобы ничего не слетело

    def update_list_of_material_types(self):
        """
        Обновляет список типов материала
        Должен вызываться в момент добавления или удаления типов материала
        :return:
        """
        self.types_of_items = self.profile.get_all_types()
        # TODO Продумать, как добавить типы в комбобоксы, и чтобы ничего не слетело

    def add_material_to_receipt(self, material: Material, component: str) -> None:
        """
        Добавляет материал в рецептуру
        :param material: Добавляемый материал
        :param component: Рецептура А или Б
        """
        if component == "A":
            self.receipt_a.add_material(material)
        elif component == "B":
            self.receipt_b.add_material(material)

    @staticmethod
    def change_type_name_material(
            material: Material,
            material_type_combobox: QComboBox,
            material_combobox: QComboBox,
    ) -> callable:
        """
        Функция, которая вызывается при смене индекса QComboBox с названием материала
        Обёрнута в замыкание для вызова триггером
        :param material_combobox: QComboBox из строки рецептуры
        :param material_type_combobox: QComboBox из строки рецептуры
        :param material: Объект Material
        :return: None
        """

        def wrapper():
            material.set_type_and_name(
                material_type_combobox.currentText(), material_combobox.currentIndex()
            )

        return wrapper

    def change_percent_material(self, material: Material, percent_line: QLineEdit) -> callable:
        """
        Меняет процент в материале, который хранится в рецептуре
        :param material:
        :param percent_line:
        :return:
        """

        def wrapper():
            material.percent = float(percent_line.text())
            self.update_mass_in_component_lines(material.receipt.component)

        return wrapper

    # Нормирует рецептуру
    def normalise_func(self, component: str) -> callable:
        """
        Нормирует рецептуру
        Обёрнута в замыкание для вызова триггером
        :param component: Рецептура А или Б
        """
        if component == "A":
            material_lines = self.material_list_a
            lock_checkbox = self.lock_checkboxies_a
            percent_lines = self.material_percent_lines_a
        elif component == "B":
            material_lines = self.material_list_b
            lock_checkbox = self.lock_checkboxies_b
            percent_lines = self.material_percent_lines_b

        def wrap():
            sum_all = 0
            sum_all_after = 0
            sum_all_without_last = 0
            total_sum_left = 100
            for material, checkbox in zip(material_lines, lock_checkbox):

                if checkbox.isChecked():
                    total_sum_left -= material
                    continue
                sum_all += material
            if sum_all:
                for material, checkbox, percent_line in zip(
                        material_lines, lock_checkbox, percent_lines
                ):
                    if checkbox.isChecked():
                        continue
                    percent = round(float(material) / sum_all * total_sum_left, 2)
                    percent_line.setText(f"{percent:.{2}f}")
                    material.percent = percent
                    sum_all_after += percent
                    if material is material_lines[-1]:
                        break
                    sum_all_without_last += percent
                if sum_all_after != 100:
                    for material, checkbox, percent_line in reversed(
                            list(zip(material_lines, lock_checkbox, percent_lines))
                    ):
                        current_numb = float(material)
                        if current_numb != 0 and not checkbox.isChecked():
                            percent = round(
                                current_numb + (total_sum_left - sum_all_after), 2
                            )
                            percent_line.setText(f"{percent:.{2}f}")
                            material.percent = percent
                            break

            if component == "A":
                self.receipt_a.set_sum_to_qt()
            elif component == "B":
                self.receipt_b.set_sum_to_qt()
            self.update_mass_in_component_lines(component)

        return wrap

    def disable_receipt(self, component: str) -> None:
        """
        Делает рецептуру неактивной на время показа окна редактирования рецептуры
        :param component: Рецептура А или Б
        """
        if component == "A":
            self.add_A_but.setEnabled(False)
            self.del_A_but.setEnabled(False)
            for i in range(len(self.material_comboboxes_a)):
                self.material_comboboxes_a[i].setEnabled(False)
                self.material_percent_lines_a[i].setEnabled(False)
                self.material_a_types[i].setEnabled(False)
                self.lock_checkboxies_a[i].setEnabled(False)
                self.normalise_A.setEnabled(False)

        elif component == "B":
            self.add_B_but.setEnabled(False)
            self.del_B_but.setEnabled(False)
            for i in range(len(self.material_comboboxes_b)):
                self.material_comboboxes_b[i].setEnabled(False)
                self.material_percent_lines_b[i].setEnabled(False)
                self.material_b_types[i].setEnabled(False)
                self.lock_checkboxies_b[i].setEnabled(False)
                self.normalise_B.setEnabled(False)

    def enable_receipt(self, component: str) -> None:
        """
        Делает рецептуру активной после закрытия окна редактирования рецептуры
        :param component: Рецептура А или Б
        """
        if component == "A":
            self.add_A_but.setEnabled(True)
            self.del_A_but.setEnabled(True)
            for i in range(len(self.material_comboboxes_a)):
                self.material_comboboxes_a[i].setEnabled(True)
                self.material_percent_lines_a[i].setEnabled(True)
                self.material_a_types[i].setEnabled(True)
                self.lock_checkboxies_a[i].setEnabled(True)
                self.normalise_A.setEnabled(True)

        elif component == "B":
            self.add_B_but.setEnabled(True)
            self.del_B_but.setEnabled(True)
            for i in range(len(self.material_comboboxes_b)):
                self.material_comboboxes_b[i].setEnabled(True)
                self.material_percent_lines_b[i].setEnabled(True)
                self.material_b_types[i].setEnabled(True)
                self.lock_checkboxies_b[i].setEnabled(True)
                self.normalise_B.setEnabled(True)

    def set_percents_from_receipt_window(self, component: str, percents: List[float]):
        """
        Устанавливает проценты из окна редактирования рецептуры.
        scope_trigger позволяет подождать, когда все значения передадутся,
        чтобы лишние разы не рассчитывать все параметры
        :param component: Рецептура А или Б
        :param percents: Список процентов для установки
        """
        if component == "A":
            material_percent_lines = self.material_percent_lines_a
            material_lines = self.material_list_a
            receipt = self.receipt_a
        elif component == "B":
            material_percent_lines = self.material_percent_lines_b
            material_lines = self.material_list_b
            receipt = self.receipt_b
        else:
            return None
        receipt.scope_trigger = len(percents) - 1
        for material, percent_line, percent in zip(
                material_lines, material_percent_lines, percents
        ):
            percent_line.setText(str(percent))
            material.percent = percent
        self.update_mass_in_component_lines(component)

    def change_receipt_color(self, component: str, color_red: bool):
        if component == "A":
            if color_red:
                style = self.style_red_but
            else:
                style = self.style
            self.a_receipt_but.setStyleSheet(style)
        elif component == "B":
            if color_red:
                style = self.style_red_but
            else:
                style = self.style
            self.b_receipt_but.setStyleSheet(style)

    def set_inner_style(self):
        """
        Устанавливает элементы стиля присуще только этому окну
        :return:
        """

        self.change_receipt_color("A", color_red=True)
        self.change_receipt_color("B", color_red=True)

        pixmap = QPixmap("icons/lock.png")
        self.label_lock_a.setPixmap(pixmap)
        self.label_lock_b.setPixmap(pixmap)

        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("icons/update.png"))
        self.update_but.setIcon(icon)

    def create_warring(self):

        rect = MyQLabel("test")
        if randint(1, 3) == 1:
            rect.extra_cor = False
        elif randint(1, 3) == 3:
            rect.extra_cor = False
            rect.base_cor = False
        self.warring_grid.add_label(rect)

    def save_receipt(self, component: str):
        """

        :param component:
        :return:
        """

        if component == "A":
            materials = self.material_list_a
            name = self.lineEdit_name_a.text()
            if not (materials and name):
                return None
            comment = self.comment_textEdit_a.toPlainText()
            mass = float(self.final_mass_a.text())

        elif component == "B":
            materials = self.material_list_b
            name = self.lineEdit_name_b.text()
            if not (materials and name):
                return None
            comment = self.comment_textEdit_b.toPlainText()
            mass = float(self.final_mass_b.text())
        else:
            return None

        receipt = self.profile.orm_db.save_receipt(materials, name, comment, mass, self.profile)
        if receipt is not None:
            self.saved_receipt_listWidget.insertItem(0, name)
            self.saved_receipt_listWidget.item(0).setToolTip(str(receipt.comment))
            self.all_receipts.insert(0, receipt)

        # self.saved_receipt_listWidget.item(0).hint

    def create_receipt(self, component: str):
        if component == "A":
            materials = self.material_list_a
            name = self.lineEdit_name_a.text()
            if not materials:
                # TODO Реализовать уведомление, что рецептура без компонентов
                return None
            comment = self.comment_textEdit_a.toPlainText()
            mass = float(self.final_mass_a.text())

        elif component == "B":
            materials = self.material_list_b
            name = self.lineEdit_name_b.text()
            if not materials:
                # TODO Реализовать уведомление, что рецептура без компонентов
                return None
            comment = self.comment_textEdit_b.toPlainText()
            mass = float(self.final_mass_b.text())
        else:
            return None

        receipt = ReceiptData(name, comment, self.profile, [mat.data_material.db_id for mat in materials],
                              [mat.percent for mat in materials], mass, datetime.now(), None,
                              [mat.data_material for mat in materials])
        return receipt

    def load_receipt(self, component: str, line: int):
        self.set_receipt(component, self.all_receipts[line])

    def set_receipt(self, component: str, receipt: ReceiptData):

        if component == 'A':
            material_list = self.material_list_a
            material_types = self.material_a_types
            material_comboboxes = self.material_comboboxes_a
            percent_lines = self.material_percent_lines_a
            self.remember_mass_a = receipt.mass
            name_line = self.lineEdit_name_a
            comment_line = self.comment_textEdit_a

        elif component == 'B':

            material_list = self.material_list_b
            material_types = self.material_b_types
            material_comboboxes = self.material_comboboxes_b
            percent_lines = self.material_percent_lines_b
            self.remember_mass_b = receipt.mass
            name_line = self.lineEdit_name_b
            comment_line = self.comment_textEdit_b

        else:
            return None

        while material_list:
            self.del_line(component)

        for line, (material_id, percent) in enumerate(receipt):
            self.add_line(component)
            material = self.profile.get_material_by_db_id(material_id)
            if material is not None:
                material_types[line].setCurrentText(material.mat_type)
                material_comboboxes[line].setCurrentText(material.name)
            else:
                # TODO Уведомление, что материал удалён из БД
                ...
            percent_lines[line].setText(str(percent))
            material_list[line].percent = percent

        if component == "A":
            self.final_mass_a.setText(str(receipt.mass))
        elif component == "B":
            self.final_mass_b.setText(str(receipt.mass))

        name_line.setText(str(receipt.name))
        comment_line.setText(str(receipt.comment))

        self.to_float(component)
        self.update_mass_in_component_lines(component)

    def delete_receipt(self, index: int):

        receipt = self.all_receipts.pop(index)
        item = self.saved_receipt_listWidget.item(index)

        self.saved_receipt_listWidget.takeItem(index)
        self.profile.orm_db.remove_receipt(receipt)

    def export_to_excel(self, component: str):

        receipts_lines = []
        receipts: List[ReceiptData] = []
        if "A" in component:
            receipt = self.create_receipt('A')
            receipts_lines.append(receipt.to_excel())
            receipts.append(receipt)

        if "B" in component:
            receipt = self.create_receipt('B')
            if receipt is None:
                return None
            receipts_lines.append(receipt.to_excel(col_start=len(receipts_lines) * 6 + 1))
            receipts.append(receipt)

        wb = opx.Workbook()
        ws = wb.active

        # lines = map(lambda x: x.to_excel(), receipts)
        lines_to_add = []

        for line in itertools.zip_longest(*receipts_lines, fillvalue=['' for _ in range(6)]):
            a = list(itertools.chain(*line))
            lines_to_add.append(a)

        # row_with_ew = ['' for _ in lines_to_add[0]]

        # for col in range(1, len(lines_to_add[-1]) + 1):
        #     if col % 6 == 3:
        #         lines_to_add[-1][col-1] = f"=SUM({letter(col)}4:{letter(col)}{len(lines_to_add) - 1})"
        #     if col % 6 == 4:
        #         for row in range(4, len(lines_to_add)):
        #             lines_to_add[row-1][col - 1] = f"={letter(col-1)}{row} * {letter(col)}{len(lines_to_add)} / 100"
        #         row_with_ew[col-1] = f'=IF({letter(col+2)}{len(lines_to_add)}>0,"EEW","AHEW")'
        #     if col % 6 == 0:
        #         for row in range(4, len(lines_to_add)):
        #             lines_to_add[row-1][col - 1] = f"=IF({letter(col-1)}{row}<>0, IF({letter(col-5)}{row}=\"Amine\", - {letter(col-3)}{row}/{letter(col-1)}{row}, {letter(col-3)}{row}/{letter(col-1)}{row}), 0)"
        #         lines_to_add[-1][col - 1] = f"=SUM({letter(col)}4:{letter(col)}{len(lines_to_add) - 1})"
        #     if col % 6 == 5:
        #         row_with_ew[col-1] = f"=ABS(100/{letter(col+1)}{len(lines_to_add) })"

        # lines_to_add.append(row_with_ew)
        lines_to_add.append([])
        if component == "AB" and self.receipt_a.receipt_counter.mass_ratio is not None:
            lines_to_add.append(['Массовое соотношение:', "", self.mass_ratio_label.text().split('\n')[1],
                                 f"={receipts[0].final_ew_link}/{receipts[1].final_ew_link}",
                                 f"={receipts[1].final_ew_link}/{receipts[0].final_ew_link}"])
            lines_to_add.append(['Стеклование:', "", self.receipt_a.receipt_counter.tg])
            lines_to_add.append(['Стеклование c коррекцией:', "", self.receipt_a.receipt_counter.tg_inf])

        for line in lines_to_add:
            ws.append(line)

        for receipt in receipts:
            ws[receipt.final_ew_link].number_format = "0.00"

        if component == "AB" and self.receipt_a.receipt_counter.mass_ratio is not None:
            ws[f"C{ws.max_row - 1}"].number_format = "0.00"
            ws[f"C{ws.max_row}"].number_format = "0.00"
            ws[f"D{ws.max_row - 2}"].number_format = "0.0000"
            ws[f"E{ws.max_row - 2}"].number_format = "0.0000"

        right_border = Border(left=Side(style='thin'))
        for cell in ws['G']:
            cell.border = right_border

        for col in range(1, ws.max_column + 1):

            if col % 6 == 1:
                ws.column_dimensions[letter(col)].width = 12
            if col % 6 == 2:
                ws.column_dimensions[letter(col)].width = 15
            if col % 6 == 3:
                ws.column_dimensions[letter(col)].width = 15
            if col % 6 == 4:
                ws.column_dimensions[letter(col)].width = 13
            if col % 6 == 5:
                ws.column_dimensions[letter(col)].width = 8
            if col % 6 == 0:
                ws.column_dimensions[letter(col)].hidden = True

        filename_base = "_".join([receipt.name if receipt.name else "noname" for receipt in receipts])
        filename = filename_base
        saved = False
        iteration = 1
        while not saved:
            try:
                wb.save(filename + '.xlsx')
                # TODO Вынести выбор запуска файла в настройки
                os.startfile(filename + '.xlsx')
                saved = True
            except:
                filename = filename_base + f"_{iteration}"
                iteration += 1


class SintezWindow(QtWidgets.QMainWindow, uic.loadUiType("windows/EEWAHEW.ui")[0]):
    def __init__(self, main_window: MyMainWindow, component):
        super(SintezWindow, self).__init__()
        self.setupUi(self)

        self.main_window = main_window
        self.component = component

        self.lines_to_change = []
        self.horizontalSlider = {}
        self.line_percent = {}
        self.line_EW = {}
        self.line_name_of_component = {}
        self.percents = []
        self.previousPercents = []
        self.sumpercent = 0
        self.checkBoxAHEW = {}
        self.checkBoxEEW = {}
        self.label_activity = {}
        self.checkBoxChange = {}

        self.__EW = 0
        self.slider_is_pushed = {}

        self.gridLayout.addItem(QSpacerItem(1, 1), 1000, 0, 1000, 5)

        # self.total_EW_lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        # self.total_EW_lineEdit.setGeometry(QtCore.QRect(100, 60, 200, 20))
        # self.total_EW_lineEdit.setObjectName("total_EW_lineEdit")

        self.material_types = []
        self.material_comboboxes = []
        self.material_percent_lines = []

        # Итератор индексов строк для алгоритма без сохранения процентов.
        self.next_line_up = self.cycle_lines_to_change()
        self.next_line_down = self.cycle_lines_to_change()

        self.label.setText("Компонент " + component)

        if component == "A":
            self.setWindowTitle("Редактирование рецептуры Компонента А")
            self.label.setText("Редактирование рецептуры Компонента А")
            self.main_window_material_comboboxes = (
                self.main_window.material_comboboxes_a
            )
            self.main_window_material_types = self.main_window.material_a_types
            self.main_window_material_percent_lines = (
                self.main_window.material_percent_lines_a
            )

        elif component == "B":
            self.setWindowTitle("Редактирование рецептуры Компонента Б")
            self.label.setText("Редактирование рецептуры Компонента Б")
            self.main_window_material_comboboxes = (
                self.main_window.material_comboboxes_b
            )
            self.main_window_material_types = self.main_window.material_b_types
            self.main_window_material_percent_lines = (
                self.main_window.material_percent_lines_b
            )
        else:
            raise TypeError

        with open("style.css", "r") as f:
            self.style, self.style_combobox, _ = f.read().split("$split$")

        self.numb_of_components = len(self.main_window_material_comboboxes)

        self.name_list = []

        self.component_ratio = None

        # Добавляем строки с компонентами
        for index, widget in enumerate(self.main_window_material_comboboxes):
            percent = float(self.main_window_material_percent_lines[index].text())
            self.percents.append(percent)
            self.previousPercents.append(percent)
            name = widget.currentText()
            # new
            self.name_list.append(name)

            self.add_line(
                index,
                self.main_window_material_types[index].currentText(),
                percent,
            )

        set_qt_stile("style.css", self)

        self.resize(680, 85 + 38 * len(self.material_types))

        self.change_font()
        # self.line = QtWidgets.QFrame(self.centralwidget)
        # self.line.setGeometry(QtCore.QRect(150, 100, 20, 300))
        # self.line.setFrameShape(QtWidgets.QFrame.VLine)
        # self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        # self.line.setObjectName("line")
        # self.line_2 = QtWidgets.QFrame(self.centralwidget)
        # self.line_2.setGeometry(QtCore.QRect(20, 140, 761, 16))
        # self.line_2.setFrameShape(QtWidgets.QFrame.HLine)
        # self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        # self.line_2.setObjectName("line_2")

    def change_font(self):
        """
        Меняет размер шрифта для соответствия в основном окне
        """
        font = QtGui.QFont("Times New Roman", self.main_window.font_size)
        big_bold_font = QtGui.QFont("Times New Roman", self.main_window.font_size_big)
        # big_bold_font.setBold(True)
        self.label.setFont(big_bold_font)
        for line_numb in range(len(self.material_types)):
            self.material_types[line_numb].setFont(font)
            self.material_comboboxes[line_numb].setFont(font)
            self.material_percent_lines[line_numb].setFont(font)

    @property
    def EW(self):
        return self.__EW

    @EW.setter
    def EW(self, value):
        # if value > 0:
        #     self.total_EW_lineEdit.setText("EEW  " + str(round(value, 2)))
        # elif value == 0:
        #     self.total_EW_lineEdit.setText("No EW")
        # else:
        #     self.total_EW_lineEdit.setText("AHEW  " + str(-round(value, 2)))
        self.__EW = value

    def checkbox_changed(self) -> None:
        """
        Вызывается при смене компонетов для изменения для расчёта нового соотношения
        """
        self.set_lines_to_change()
        self.set_base_ratio()
        self.next_line_up = self.cycle_lines_to_change()
        self.next_line_down = self.cycle_lines_to_change()

    def set_base_ratio(self):
        self.component_ratio = {}
        # Составим словарь соотношений
        for index, name in enumerate(self.name_list):
            if len(self.name_list) > index + 1:
                for next_index, next_name in enumerate(
                        self.name_list[index + 1:], start=index + 1
                ):
                    if self.percents[next_index] != 0:
                        self.component_ratio[(index, next_index)] = (
                                self.percents[index] / self.percents[next_index]
                        )
                    else:
                        self.component_ratio[(index, next_index)] = inf

    def slider_is_moved(self, numb_of_slider):
        def wrapper():
            value_of_slider = int(self.horizontalSlider[numb_of_slider].value()) / 100
            # self.try_to_change(numb_of_slider, self.percents[numb_of_slider], value_of_slider)
            self.percents[numb_of_slider] = value_of_slider
            # self.count_EW()
            self.line_percent[numb_of_slider].setText(str(value_of_slider))

        return wrapper

    def change_percent_in_line(self, numb_of_line):
        def wrapper():
            text = self.line_percent[numb_of_line].text()
            if text.isdigit():
                numb = float(text)
                self.horizontalSlider[numb_of_line].setSliderPosition(numb * 100)
                self.line_percent[numb_of_line].setText(text)
            else:
                # Вписать инструкцию, которая будет сообщать об ошибке
                pass

        return wrapper

    def add_line(
            self,
            numb_of_line,
            mat_type,
            percent=None,
    ):

        items_type = self.material_types
        items = self.material_comboboxes
        items_lines = self.material_percent_lines
        grid = self.gridLayout
        material_combobox = QComboBox()
        materia_type_combobox = QComboBox()

        materia_type_combobox.setStyleSheet(self.style_combobox)
        material_combobox.setStyleSheet(self.style_combobox)
        # Подтянуть соответствующий индекс
        materia_type_combobox.addItems(self.main_window.types_of_items)
        materia_type_combobox.setCurrentIndex(
            self.main_window_material_types[numb_of_line].currentIndex()
        )

        materia_type_combobox.setFixedWidth(60)

        materia_type_combobox.currentIndexChanged.connect(
            self.main_window.change_list_of_materials(
                material_combobox, materia_type_combobox
            )
        )

        # Подцепить соответствующие вещества
        # TODO Переделать комбобоксы на строки без права смены материала
        material_combobox.addItems(self.main_window.list_of_material_names[mat_type])
        material_combobox.setCurrentIndex(
            self.main_window_material_comboboxes[numb_of_line].currentIndex()
        )
        material_combobox.setFixedWidth(120)

        line = QLineEdit()
        line.setText(self.main_window_material_percent_lines[numb_of_line].text())
        # line.editingFinished.connect(lambda: self.to_float(component))
        # line.editingFinished.connect(lambda: self.count_sum(component))

        items_type.append(materia_type_combobox)
        items.append(material_combobox)
        items_lines.append(line)
        row_count = numb_of_line
        grid.addWidget(materia_type_combobox, row_count + 1, 0)
        grid.addWidget(material_combobox, row_count + 1, 1)
        grid.addWidget(line, row_count + 1, 2)

        # grid.addWidget(final_label, row_count + 2, 1, alignment=QtCore.Qt.AlignRight)
        # grid.addWidget(final_label_numb, row_count + 2, 2)
        # self.count_sum(component)
        line.setFixedWidth(60)
        self.line_percent[numb_of_line] = line
        # self.line_name_of_component[numb_of_line] = QComboBox(self.centralwidget)
        # self.line_name_of_component[numb_of_line].setGeometry(QtCore.QRect(x + 140, 110 + interval * numb_of_line, 141, 20))
        # self.line_name_of_component[numb_of_line].setObjectName(f"line_name_of_component{numb_of_line}")

        # Создаём окно для процентов
        # self.line_percent[numb_of_line] = QtWidgets.QLineEdit(self.centralwidget)
        # self.line_percent[numb_of_line].setGeometry(QtCore.QRect(x + 300, 110 + interval * numb_of_line, 51, 20))
        # self.line_percent[numb_of_line].setObjectName(f"line_percent{numb_of_line}")
        if percent:
            self.line_percent[numb_of_line].setText(str(percent))
        self.line_percent[numb_of_line].editingFinished.connect(
            self.try_to_change(numb_of_line, "line")
        )

        self.checkBoxChange[numb_of_line] = QtWidgets.QCheckBox()
        # self.checkBoxChange[numb_of_line].setGeometry(QtCore.QRect(x + 370, 112 + interval * numb_of_line, 16, 17))
        # self.checkBoxChange[numb_of_line].setText("")
        # self.checkBoxChange[numb_of_line].setObjectName(f"checkBoxChange{numb_of_line}")
        self.checkBoxChange[numb_of_line].clicked.connect(self.checkbox_changed)
        grid.addWidget(self.checkBoxChange[numb_of_line], row_count + 1, 3)

        slider = QtWidgets.QSlider()
        # self.horizontalSlider[numb_of_line].setGeometry(QtCore.QRect(x + 390, 110 + interval * numb_of_line, 400, 20))
        # self.horizontalSlider[numb_of_line].setProperty("value", 20)
        # self.horizontalSlider[numb_of_line].setSliderPosition(10 + 3 * numb_of_line)
        slider.setOrientation(QtCore.Qt.Horizontal)
        slider.setTickInterval(10)
        slider.setObjectName(f"horizontalSlider{numb_of_line}")
        slider.setRange(0, 10000)

        self.horizontalSlider[numb_of_line] = slider
        grid.addWidget(slider, row_count + 1, 4)
        if percent:
            slider.setSliderPosition(percent * 100)

        slider.valueChanged.connect(self.try_to_change(numb_of_line, "slider"))
        # self.horizontalSlider[numb_of_line].sliderMoved.connect(self.try_to_change(numb_of_line, 'slider'))
        # slider.sliderPressed.connect(self.slider_push_changer(numb_of_line, True))
        slider.sliderReleased.connect(self.set_percents)
        # self.slider_is_pushed[numb_of_line] = False

    def slider_push_changer(self, line: int, is_push: bool):
        def wrapper():
            self.slider_is_pushed[line] = is_push

        return wrapper

    def try_to_change_dont_save_ratio(self, numb_of_line, source):
        def wrapper():

            # Фиксируем слайдер, если пытаются двигать отмеченный
            if self.checkBoxChange[numb_of_line].isChecked():
                self.horizontalSlider[numb_of_line].setSliderPosition(
                    self.percents[numb_of_line] * 100
                )
                return None

            lines_to_change = []
            for line in range(self.numb_of_components):
                if self.checkBoxChange[line].isChecked() and line != numb_of_line:
                    lines_to_change.append(line)
            # Фиксируем слайдер, если ни один слайдер не выбран
            if len(lines_to_change) == 0:
                self.horizontalSlider[numb_of_line].setSliderPosition(
                    self.percents[numb_of_line] * 100
                )

            previos_value = self.previousPercents[numb_of_line]

            if source == "slider":
                new_value = round(
                    int(self.horizontalSlider[numb_of_line].value()) / 100, 2
                )
            else:
                new_value = self.line_percent[numb_of_line].text()
                try:
                    new_value = round(float(new_value), 2)
                except:
                    self.line_percent[numb_of_line].setText(
                        str(self.percents[numb_of_line])
                    )
                    return None

            delta = round(new_value - previos_value, 2)

            if delta > 0:
                change_way_is_up = True
            else:
                change_way_is_up = False
                delta = -delta

            if delta < 0.01:
                return None

            for line in range(self.numb_of_components):
                self.previousPercents[line] = self.percents[line]

            # Функция, меняющая компоненты без сохранения EW

            sum_percent = 0
            sum_ostatok_percent = 0

            for line in lines_to_change:
                sum_percent += self.percents[line]
                sum_ostatok_percent += 100 - self.percents[line]

            if ((sum_percent > delta) and change_way_is_up) or (
                    (sum_ostatok_percent > delta) and not change_way_is_up
            ):

                self.percents[numb_of_line] = round(self.percents[numb_of_line], 2)
                break_flag = []

                if change_way_is_up:
                    cycle_lines = self.next_line_up
                else:
                    cycle_lines = self.next_line_down

                for line in cycle_lines:
                    # Ходим по концентрациям других продуктов и меняем при проходе на 0,01%, если там что-то еще осталось
                    # Когда ничего не осталось, возвращаем компонент, который меняли обратно на оставшуюся дельту

                    if delta < 0.01:
                        break

                    if not change_way_is_up:
                        self.percents[line] += 0.01
                        self.percents[line] = round(self.percents[line], 2)
                        self.percents[numb_of_line] -= 0.01
                        self.percents[numb_of_line] = round(
                            self.percents[numb_of_line], 2
                        )

                    if self.percents[line] == 0:
                        if line not in break_flag:
                            break_flag.append(line)
                        if len(break_flag) == len(lines_to_change):
                            break
                        continue

                    if change_way_is_up:
                        self.percents[line] -= 0.01
                        self.percents[line] = round(self.percents[line], 2)
                        self.percents[numb_of_line] += 0.01
                        self.percents[numb_of_line] = round(
                            self.percents[numb_of_line], 2
                        )

                    delta -= 0.01
                    delta = round(delta, 2)

            else:
                # Когда нам надо добить до конца
                if change_way_is_up:
                    for line in lines_to_change:
                        self.percents[numb_of_line] += self.percents[line]
                        self.percents[numb_of_line] = round(
                            self.percents[numb_of_line], 2
                        )
                        self.percents[line] = 0
                else:
                    for line in lines_to_change:
                        self.percents[numb_of_line] -= 100 - self.percents[line]
                        self.percents[numb_of_line] = round(
                            self.percents[numb_of_line], 2
                        )
                        self.percents[line] = 100

            self.set_percents(numb_of_line)
            sum_percent_all = 0
            for i in self.percents:
                sum_percent_all += i

            for line in range(self.numb_of_components):
                self.previousPercents[line] = self.percents[line]

            self.main_window.set_percents_from_receipt_window(
                self.component,
                [self.percents[i] for i in range(self.numb_of_components)],
            )

        return wrapper

    def try_to_change(self, numb_of_line, source):
        def wrapper():
            # Фиксируем слайдер, если пытаются двигать отмеченный
            if self.checkBoxChange[numb_of_line].isChecked():
                self.horizontalSlider[numb_of_line].setSliderPosition(
                    self.percents[numb_of_line] * 100
                )
                return None

            # Фиксируем слайдер, если ни один слайдер не выбран
            if len(self.lines_to_change) == 0:
                self.horizontalSlider[numb_of_line].setSliderPosition(
                    self.percents[numb_of_line] * 100
                )
                return None

            lines_to_change = self.lines_to_change.copy()
            previos_value = self.previousPercents[numb_of_line]

            if source == "slider":
                new_value = round(
                    int(self.horizontalSlider[numb_of_line].value()) / 100, 2
                )
            else:
                new_value = self.line_percent[numb_of_line].text()
                try:
                    new_value = round(float(new_value), 2)
                except:
                    self.line_percent[numb_of_line].setText(
                        str(self.percents[numb_of_line])
                    )
                    return None

            delta = round(new_value - previos_value, 2)

            if delta > 0:
                # Ползунок, который дёргаем идёт вверх
                change_way_is_up = True
            else:
                # Ползунок, который дёргаем идёт вниз
                change_way_is_up = False
                delta = -delta

            # ======= Если только один слайдер в пару, то только его и меняем
            if len(self.lines_to_change) == 1:
                # Просто смотрим дельту и ставим проценты по условиям
                changed_line = lines_to_change[0]
                if change_way_is_up:
                    if self.percents[changed_line] >= delta:
                        self.percents[changed_line] = (
                                self.percents[changed_line] - delta
                        )
                        self.percents[numb_of_line] = (
                                self.percents[numb_of_line] + delta
                        )
                    else:
                        self.percents[numb_of_line] = 100.0 - sum(
                            [
                                p
                                for index, p in enumerate(self.percents)
                                if index not in [changed_line, numb_of_line]
                            ]
                        )
                        self.percents[changed_line] = 0.0
                        self.horizontalSlider[numb_of_line].setSliderPosition(
                            self.percents[numb_of_line] * 100
                        )
                else:
                    if 100 - self.percents[changed_line] > delta:
                        self.percents[changed_line] = (
                                self.percents[changed_line] + delta
                        )
                        self.percents[numb_of_line] = (
                                self.percents[numb_of_line] - delta
                        )
                    else:
                        self.percents[changed_line] = 100.0 - sum(
                            [
                                p
                                for index, p in enumerate(self.percents)
                                if index not in [changed_line, numb_of_line]
                            ]
                        )
                        self.horizontalSlider[numb_of_line].setSliderPosition(
                            self.percents[numb_of_line] * 100
                        )
                        self.percents[numb_of_line] = 0.0
                self.percents[numb_of_line] = round(self.percents[numb_of_line], 2)
                self.percents[changed_line] = round(self.percents[changed_line], 2)
                for line in range(self.numb_of_components):
                    self.previousPercents[line] = self.percents[line]
                self.set_percents(numb_of_line)
                return None

            # ======= Логика с сохранением соотношений
            if self.component_ratio is None:
                self.set_base_ratio()
            current_base_sootnoshenie = self.component_ratio.copy()

            all_keys = set(current_base_sootnoshenie.keys())
            used_keys = set()

            for first, second in current_base_sootnoshenie.keys():
                if first in self.lines_to_change and second in self.lines_to_change:
                    used_keys.add((first, second))

            for index in all_keys - used_keys:
                del current_base_sootnoshenie[index]
            current_percent_line = self.percents.copy()

            fix_loop = 0
            while delta >= 0.01:
                step_limit = min(
                    [
                        percent
                        for row, percent in enumerate(current_percent_line)
                        if row in self.lines_to_change
                    ]
                    + [delta]
                )
                if step_limit > 20:
                    step = 5
                elif step_limit > 2:
                    step = 1
                elif step_limit > 0.2:
                    step = 0.1
                else:
                    step = 0.01
                current_percent_ratio = self.get_percent_ratio(current_percent_line)
                if fix_loop == 2:
                    break

                index = self.choose_component_to_change(
                    current_percent_ratio,
                    self.component_ratio,
                    self.lines_to_change,
                    change_way_is_up,
                )
                if change_way_is_up:
                    if current_percent_line[index] - step >= 0:
                        current_percent_line[numb_of_line] += step
                        current_percent_line[numb_of_line] = round(
                            current_percent_line[numb_of_line], 2
                        )
                        current_percent_line[index] -= step
                        current_percent_line[index] = round(
                            current_percent_line[index], 2
                        )
                    else:
                        step = current_percent_line[index]
                        current_percent_line[numb_of_line] += step
                        current_percent_line[numb_of_line] = round(
                            current_percent_line[numb_of_line], 2
                        )
                        current_percent_line[index] -= step
                        current_percent_line[index] = round(
                            current_percent_line[index], 2
                        )

                else:
                    if current_percent_line[index] - step <= 100:

                        current_percent_line[numb_of_line] -= step
                        current_percent_line[numb_of_line] = round(
                            current_percent_line[numb_of_line], 2
                        )
                        current_percent_line[index] += step
                        current_percent_line[index] = round(
                            current_percent_line[index], 2
                        )
                    else:
                        step = current_percent_line[index]
                        current_percent_line[numb_of_line] -= step
                        current_percent_line[numb_of_line] = round(
                            current_percent_line[numb_of_line], 2
                        )
                        current_percent_line[index] += step
                        current_percent_line[index] = round(
                            current_percent_line[index], 2
                        )

                if self.previousPercents == self.percents:
                    fix_loop += 1

                delta -= step
                delta = round(delta, 2)
                for line in range(self.numb_of_components):
                    self.previousPercents[line] = self.percents[line]
                self.percents = current_percent_line

            self.set_percents(numb_of_line)
            self.horizontalSlider[numb_of_line].setSliderPosition(
                self.percents[numb_of_line] * 100
            )

        return wrapper

    @staticmethod
    def choose_component_to_change(
            current_ratio, base_ratio, lines_to_change, change_way_is_up
    ):
        if len(lines_to_change) == 1:
            return lines_to_change[0]
        ds = dict()
        for pair in current_ratio.keys():
            if current_ratio[pair] == 0:
                ds[pair] = inf
            else:
                ds[pair] = base_ratio[pair] / current_ratio[pair] - 1
        biggest_change = 0
        current_pair = None
        for pair, value in ds.items():
            if pair[0] in lines_to_change and pair[1] in lines_to_change:
                if abs(value) > biggest_change:
                    biggest_change = abs(value)
                    current_pair = pair
        if current_pair is None:
            return lines_to_change[0]
        if change_way_is_up:
            if ds[current_pair] > 0:
                return current_pair[1]
            else:
                return current_pair[0]
        else:
            if ds[current_pair] > 0:
                return current_pair[0]
            else:
                return current_pair[1]

    def get_percent_ratio(self, percent_list):
        len_percent_list = len(percent_list)
        ds = {}
        for index, percent in enumerate(percent_list):
            if len_percent_list > index + 1:
                for next_index, next_percent in enumerate(
                        percent_list[index + 1:], start=index + 1
                ):
                    if next_percent != 0:
                        ds[(index, next_index)] = percent / next_percent
                    else:
                        ds[(index, next_index)] = inf
        return ds

    @staticmethod
    def get_pair_to_change(base_ds, changed_ds):
        ds_delta = changed_ds / base_ds - 1

        return max(*ds_delta.items(), key=lambda x: abs(x[1]))

    def set_percents(self, current_line=-1):
        """
        Устанавливает слайдеры и проценты в SintezWindow и MyMainWindow
        :param current_line:
        :return:
        """
        for line in range(self.numb_of_components):
            self.line_percent[line].setText(str(self.percents[line]))
            if line == current_line:
                continue
            self.horizontalSlider[line].setSliderPosition(self.percents[line] * 100)
        self.main_window.set_percents_from_receipt_window(self.component, self.percents)

    def cycle_lines_to_change(self) -> Union[Iterable, callable]:
        """
        Функция для хождения по строкам при изменении процентов без сохранения соотношения
        Необходимо получить итератор и ходить по нему. Менять итератор на новый при смене checkbox!
        """
        for row in cycle(self.lines_to_change):
            yield row

    def set_lines_to_change(self):
        self.lines_to_change = []
        for line in self.checkBoxChange:
            if self.checkBoxChange[line].isChecked():
                self.lines_to_change.append(line)

    def closeEvent(self, a0: QtGui.QCloseEvent) -> None:
        if self.component == "A":
            self.main_window.a_receipt_window = None
        if self.component == "B":
            self.main_window.b_receipt_window = None

        self.main_window.enable_receipt(self.component)
        self.close()


class PairReactWindow(
    QtWidgets.QMainWindow, uic.loadUiType("windows/choose_pair_react.ui")[0]
):
    def __init__(
            self, main_window: MyMainWindow, receipt_a: Receipt, receipt_b: Receipt
    ):
        super(PairReactWindow, self).__init__()
        self.setupUi(self)

        set_qt_stile("style.css", self)

        self.main_window = main_window
        self.receipt_a = receipt_a
        self.receipt_b = receipt_b
        self.receipt_a.pair_react_window = self
        self.receipt_b.pair_react_window = self

        self.labels_a = []
        self.labels_b = []
        self.checkboxes_a = []
        self.checkboxes_b = []
        self.checkboxes_a_means = defaultdict(lambda: True)
        self.checkboxes_b_means = defaultdict(lambda: True)
        self.pairs_to_react_a = self.receipt_a.react_pairs
        self.pairs_to_react_b = self.receipt_a.react_pairs

        self.gridLayout_a.addItem(QSpacerItem(100, 10), 100, 0, 100, 2)
        self.gridLayout_b.addItem(QSpacerItem(100, 10), 100, 0, 100, 2)

        self.update_component("A")
        self.update_component("B")

    def update_component(self, component):
        if component == "A":
            while self.labels_a:
                self.labels_a.pop(0).deleteLater()
                self.checkboxes_a.pop(0).deleteLater()
            for pair in self.receipt_a.all_pairs_material:
                self.add_line(
                    pair,
                    self.gridLayout_a,
                    self.labels_a,
                    self.checkboxes_a,
                    self.checkboxes_a_means,
                )
        elif component == "B":
            while self.labels_b:
                self.labels_b.pop(0).deleteLater()
                self.checkboxes_b.pop(0).deleteLater()
            for pair in self.receipt_b.all_pairs_material:
                self.add_line(
                    pair,
                    self.gridLayout_b,
                    self.labels_b,
                    self.checkboxes_b,
                    self.checkboxes_b_means,
                )

    def add_line(
            self,
            pair: Tuple[Material],
            layout: QGridLayout,
            labels_list: list,
            checkboxes_list: list,
            checkboxes_means: defaultdict,
    ):
        label = QLabel()
        label.setText(f"{pair[0]} + {pair[1]}")
        labels_list.append(label)

        checkbox = QCheckBox()
        checkbox.setChecked(checkboxes_means[f"{pair[0]}-{pair[1]}"])
        checkbox.setFixedWidth(20)
        checkbox.setFixedHeight(20)

        checkbox.stateChanged.connect(
            self.change_checkbox_state(checkbox, checkboxes_means, pair)
        )
        # TODO Добавить функцию пересчёта при смене пары
        checkboxes_list.append(checkbox)

        row_count = layout.count()
        layout.addWidget(checkbox, row_count + 1, 0)
        layout.addWidget(label, row_count + 1, 1)

    def change_checkbox_state(
            self, checkbox: QCheckBox, checkboxes_means: defaultdict, pair: tuple
    ):
        def wrapper():
            checkboxes_means[f"{pair[0]}-{pair[1]}"] = checkbox.isChecked()
            self.get_react_pairs(pair[0].receipt.component)

        return wrapper

    def get_react_pairs(self, component):
        if component == "A":
            checkboxes_list = self.checkboxes_a
            all_pairs = self.receipt_a.all_pairs_material
            pairs_to_react = copy(self.pairs_to_react_a)
        elif component == "B":
            checkboxes_list = self.checkboxes_b
            all_pairs = self.receipt_b.all_pairs_material
            pairs_to_react = copy(self.pairs_to_react_b)
        else:
            return None

        while pairs_to_react:
            pairs_to_react.pop(0)

        for checkbox, pair in zip(checkboxes_list, all_pairs):
            if checkbox.isChecked():
                pairs_to_react.append(pair)
        return pairs_to_react

    def closeEvent(self, a0: QtGui.QCloseEvent) -> None:
        if not all(
                chbox.isChecked() for chbox in self.checkboxes_b + self.checkboxes_a
        ):
            self.main_window.sintez_pair_label.setText("Ступенчатый синтез")
        a0.accept()


class ProfileManagerWindow(
    QtWidgets.QMainWindow, uic.loadUiType("windows/profile_manager_window.ui")[0]
):
    add_profile_but: QPushButton
    remove_profile_but: QPushButton
    prof_name_line: QLineEdit
    profile_widget: QListWidget
    remember_profile_checkBox: QCheckBox

    def __init__(self, profile_list: list, init_class):
        super(ProfileManagerWindow, self).__init__()
        self.setupUi(self)
        self.profile_list = profile_list
        self.init_class = init_class
        self.update_profiles()

        self.choose_profile_but.clicked.connect(self.choose_profile)
        self.edit_profile_but.clicked.connect(self.edit_materials)
        self.add_profile_but.clicked.connect(self.add_profile)
        self.remove_profile_but.clicked.connect(self.remove_profile)

        # Вынести путь к стилю в настройки
        set_qt_stile("style.css", self)

        self.edit_material_window = None
        self.main_window = None
        self.profile_widget.setCurrentRow(0)

    def choose_profile(self):
        # TODO Реализовать логику по подключению данных профиля
        prof_name = self.profile_list[self.profile_widget.currentIndex().row()]
        if self.remember_profile_checkBox.isChecked():
            user_name: str = getpass.getuser()
            computer_name: str = socket.gethostname()
            self.init_class.orm_db.add_computer_to_profile(user_name, computer_name, prof_name)
            ...
        self.close()
        self.init_class.setup_program(prof_name)
        # self.main_window.show()

    def edit_materials(self):
        prof_name = self.profile_list[self.profile_widget.currentIndex().row()]

        self.edit_material_window = EditDataWindow(
            self, self.init_class.orm_db.read_profile(prof_name)
        )
        self.close()
        self.edit_material_window.show()

    def add_profile(self):
        """
        Добавление профиля в БД
        :return:
        """
        name = self.prof_name_line.text()
        if name and name not in self.profile_list:
            self.init_class.orm_db.add_profile(name)
            self.profile_list.append(name)
            self.profile_widget.addItem(name)

    def remove_profile(self):
        line = self.profile_widget.currentIndex().row()
        if line != -1:
            name = self.profile_list.pop(line)
            self.init_class.orm_db.remove_profile(name)
            self.update_profiles()

    def update_profiles(self):
        """
        Обновляем имена в виджете
        """
        self.profile_widget.clear()
        for name in self.profile_list:
            self.profile_widget.addItem(name)


class AcceptDropWidget(QLabel):
    def __init__(self, parent: MyMainWindow, action: str, component: str = None):
        super().__init__(parent)
        self.main_window = parent
        self.setAcceptDrops(True)
        self.component = component
        self.action = action

    # вызывается при попадании в область
    def dragEnterEvent(self, e):
        # Позволяет перетащить объект в этот виджет
        e.accept()
        # e.ignore()

    # вызывается, когда объект кидается в области
    def dropEvent(self, e):
        e.accept()
        text = e.mimeData().text()
        if text.isdigit():
            index = int(text)
            if self.action == 'load':
                self.main_window.load_receipt(self.component, index)
            elif self.action == 'delete':
                self.main_window.delete_receipt(index)


class SavedReceiptWidget(QListWidget):
    def __init__(self, parent: MyMainWindow):
        super().__init__(parent)
        self.main_window = parent

        self.resize(QSize(150, 520))
        self.move(QPoint(920, 60))

        self.setDragEnabled(True)

    def mimeData(self, my_list: Iterable, lwi: QListWidgetItem = None):
        mimedata = super().mimeData(my_list)
        mimedata.setText(str(self.currentIndex().row()))
        return mimedata
