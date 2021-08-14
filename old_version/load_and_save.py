import os
from itertools import zip_longest

import openpyxl as opx
from PyQt5.QtWidgets import QFileDialog
from openpyxl.styles import Alignment, Font


def create_one_komponent_rows(
    name: str,
    mat_type_list: list,
    mat_list: list,
    percent_list: list,
    ew_list: list,
    mass: float,
    component: str,
    side_left: bool,
):
    all_rows = [
        [f"Компонент {component}", "", name, "", ""],
        ["" for _ in range(5)],
        ["" for _ in range(5)],
        ["Тип", "Материал", "EW", "Содержание, %", "Навеска, г"],
    ]
    row_numb_start = 5
    row_numb = row_numb_start
    last_row = row_numb + len(mat_list)
    ew_string = "=100/("
    for mat_type, mat, ew, percent in zip(
        mat_type_list, mat_list, ew_list, percent_list
    ):
        row = [mat_type, mat, ew, float(percent)]
        if side_left:
            row.append(f"=D{row_numb}*E{last_row}/100")
            ew_string += f'+ЕСЛИ(A{row_numb}="Epoxy";D{row_numb}/C{row_numb};ЕСЛИ(A{row_numb}="Amine";-D{row_numb}/C{row_numb};0))'
        else:
            row.append(f"=J{row_numb}*K{last_row}/100")
            ew_string += f'+ЕСЛИ(G{row_numb}="Epoxy";J{row_numb}/I{row_numb};ЕСЛИ(G{row_numb}="Amine";-J{row_numb}/I{row_numb};0))'
        all_rows.append(row)
        row_numb += 1
    ew_string += ")"
    if side_left:
        total_percent = f"=SUM(D{row_numb_start}:D{last_row-1})"
    else:
        total_percent = f"=SUM(J{row_numb_start}:J{last_row - 1})"

    all_rows.append(["", "", "ИТОГО:", total_percent, mass])
    print(ew_string)
    return all_rows, ew_string


def save_receipt(
    name_a: str = None,
    name_b: str = None,
    mat_type_a_list: list = None,
    mat_type_b_list: list = None,
    mat_a_list: list = None,
    mat_b_list: list = None,
    percent_a_list: list = None,
    percent_b_list: list = None,
    ew_a_list: list = None,
    ew_b_list: list = None,
    mass_a: float = 100,
    mass_b: float = 100,
    save_a: bool = False,
    save_b: bool = False,
):

    wb = opx.Workbook()
    ws = wb.active

    if save_a and save_b:
        rows_a, ew_string_a = create_one_komponent_rows(
            name_a,
            mat_type_a_list,
            mat_a_list,
            percent_a_list,
            ew_a_list,
            mass_a,
            "А",
            True,
        )
        rows_b, ew_string_b = create_one_komponent_rows(
            name_b,
            mat_type_b_list,
            mat_b_list,
            percent_b_list,
            ew_b_list,
            mass_b,
            "Б",
            False,
        )
        if name_a and name_b:
            filename = name_a + "_" + name_b
        else:
            filename = name_a or name_b
        for row_a, row_b in zip_longest(
            rows_a, rows_b, fillvalue=["" for _ in range(5)]
        ):
            ws.append(row_a + ["|"] + row_b)
        # ws["B2"] = ew_string_a
        # ws["H2"] = ew_string_b
    elif save_a:
        rows_a, ew_string_a = create_one_komponent_rows(
            name_a,
            mat_type_a_list,
            mat_a_list,
            percent_a_list,
            ew_a_list,
            mass_a,
            "А",
            True,
        )
        for row in rows_a:
            ws.append(row)
        # ws["B2"] = ew_string_a
        filename = name_a
    elif save_b:
        rows_b, ew_string_b = create_one_komponent_rows(
            name_b,
            mat_type_b_list,
            mat_b_list,
            percent_b_list,
            ew_b_list,
            mass_b,
            "Б",
            True,
        )
        for row in rows_b:
            ws.append(row)
        # ws["B2"] = ew_string_b
        filename = name_b
    else:
        return None

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 3
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 7
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 10
    ws.merge_cells("A1:B1")
    ws.merge_cells("C1:D1")
    ws.merge_cells("G1:H1")
    ws.merge_cells("I1:J1")
    ws["C1"].font = Font(bold=True, size=16)
    ws["I1"].font = Font(bold=True, size=16)
    prev = None
    prev_prev = None
    for row in ws.iter_rows(1):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.value == "ИТОГО:":
                ws.merge_cells(f"{prev_prev.coordinate}:{cell.coordinate}")
                prev_prev.value = "ИТОГО:"
                prev_prev.alignment = Alignment(horizontal="right", vertical="center")
            prev_prev = prev
            prev = cell

    file = QFileDialog.getSaveFileName(
        None, "Сохранить синтез", filename if filename != "_" else "", "xlsx(*.xlsx)"
    )
    if file[0]:
        filename += ".xlsx"
        wb.save(file[0])

        os.startfile(file[0])
